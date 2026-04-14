"""Otonom Tartışma - Kişilikler kendi aralarında saatlerce tartışır, sonuçları kaydeder."""

import os
import json
from datetime import datetime

import ollama
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from model_config import get_chat_model
from debate import persona_manager

MODEL_NAME = get_chat_model()
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
DEBATES_DIR = os.path.join(DATA_DIR, "debates")


class AutoDebate:
    """
    Otonom tartışma: kişilikler kendi aralarında tartışır,
    fikirler üretir, notlar alır, dosyaya kaydeder.
    """

    def __init__(self, topic: str, personas: list[str], rounds: int = 20,
                 on_message=None, on_status=None):
        """
        topic: Tartışma konusu
        personas: Katılımcı kişilik listesi (2+)
        rounds: Toplam tur sayısı
        on_message: callback(persona, message, round_num)
        on_status: callback(status_text)
        """
        self.topic = topic
        self.personas = personas
        self.rounds = rounds
        self.on_message = on_message or (lambda *a: None)
        self.on_status = on_status or (lambda m: None)

        self.messages = []  # Tüm mesajlar
        self.key_ideas = []  # Önemli fikirler / çıkarımlar
        self.running = False

        os.makedirs(DEBATES_DIR, exist_ok=True)

    def stop(self):
        """Tartışmayı durdur."""
        self.running = False

    def run(self):
        """Tartışmayı başlat ve bitene kadar çalıştır."""
        self.running = True
        histories = {}

        # Her kişilik için system prompt hazırla
        for p in self.personas:
            others = [x for x in self.personas if x != p]
            sys_prompt = persona_manager.get_system_prompt(p, talking_to=others[0] if others else None)
            sys_prompt += (
                f"\n\nBir toplantıdasın. Konu: {self.topic}\n"
                f"Katılımcılar: {', '.join(self.personas)}\n"
                f"Kısa ve öz konuş (3-5 cümle). Her turda yeni bir fikir veya bakış açısı sun. "
                f"Tekrara düşme, önceki fikirleri geliştir veya karşı çık."
            )
            histories[p] = [{"role": "system", "content": sys_prompt}]

        # İlk tur: herkes konuyu açar
        self.on_status(f"Tartışma başlıyor: {' vs '.join(self.personas)} — {self.rounds} tur")

        for round_num in range(1, self.rounds + 1):
            if not self.running:
                self.on_status("Tartışma kullanıcı tarafından durduruldu.")
                break

            self.on_status(f"Tur {round_num}/{self.rounds}")

            for i, persona in enumerate(self.personas):
                if not self.running:
                    break

                # Son mesajları context olarak ekle
                recent = self.messages[-6:]  # Son 6 mesaj
                context = ""
                if recent:
                    context = "Son konuşmalar:\n"
                    for m in recent:
                        context += f"  {m['persona']}: {m['text'][:200]}\n"

                if round_num == 1 and i == 0:
                    prompt = f"Toplantı başlıyor. Konu: {self.topic}\nGörüşünü paylaş."
                else:
                    prompt = (
                        f"{context}\n"
                        f"Tur {round_num}. Kendi bakış açından katkıda bulun. "
                        f"Önceki fikirlere katıl veya karşı çık, yeni öneriler sun."
                    )

                histories[persona].append({"role": "user", "content": prompt})

                try:
                    response = ollama.chat(
                        model=MODEL_NAME,
                        messages=histories[persona][-10:],  # Son 10 mesaj (RAM tasarrufu)
                    )
                    msg = response["message"]["content"]
                except Exception as e:
                    msg = f"[Hata: {e}]"

                histories[persona].append({"role": "assistant", "content": msg})

                record = {
                    "persona": persona,
                    "text": msg,
                    "round": round_num,
                    "timestamp": datetime.now().isoformat(),
                }
                self.messages.append(record)
                self.on_message(persona, msg, round_num)

            # Her 5 turda özet çıkar
            if round_num % 5 == 0 and self.running:
                self._extract_ideas(round_num)

        # Final özet
        if self.messages:
            self._final_summary()

        # Dosyalara kaydet
        self._save_all()
        self.on_status("Tartışma tamamlandı! Dosyalar kaydedildi.")
        return self.messages

    def _extract_ideas(self, round_num: int):
        """Her 5 turda ana fikirleri çıkar."""
        recent = self.messages[-10:]
        context = "\n".join(f"{m['persona']}: {m['text'][:300]}" for m in recent)

        response = ollama.chat(
            model=MODEL_NAME,
            messages=[
                {"role": "system", "content": "Sen bir toplantı sekretersin. Türkçe yaz."},
                {"role": "user", "content": (
                    f"Tartışma konusu: {self.topic}\n\n"
                    f"Son konuşmalar:\n{context}\n\n"
                    f"Bu bölümden çıkan 3-5 ana fikri madde madde listele. "
                    f"Kısa ve öz ol."
                )},
            ],
        )
        ideas = response["message"]["content"]
        self.key_ideas.append({
            "round": round_num,
            "ideas": ideas,
            "timestamp": datetime.now().isoformat(),
        })
        self.on_message("📋 Sekreter", f"Tur {round_num} özeti:\n{ideas}", round_num)

    def _final_summary(self):
        """Tartışmanın final özetini çıkar."""
        all_ideas = "\n".join(k["ideas"] for k in self.key_ideas)
        first_msgs = self.messages[:6]
        last_msgs = self.messages[-6:]

        context = "İlk fikirler:\n"
        context += "\n".join(f"  {m['persona']}: {m['text'][:200]}" for m in first_msgs)
        context += "\n\nSon fikirler:\n"
        context += "\n".join(f"  {m['persona']}: {m['text'][:200]}" for m in last_msgs)

        if all_ideas:
            context += f"\n\nAra özetler:\n{all_ideas}"

        response = ollama.chat(
            model=MODEL_NAME,
            messages=[
                {"role": "system", "content": "Sen üst düzey bir iş stratejistisin. Türkçe yaz."},
                {"role": "user", "content": (
                    f"Konu: {self.topic}\n"
                    f"Katılımcılar: {', '.join(self.personas)}\n"
                    f"Toplam {len(self.messages)} mesaj, {self.messages[-1]['round'] if self.messages else 0} tur.\n\n"
                    f"{context}\n\n"
                    f"Tartışmanın final raporunu yaz:\n"
                    f"1. Ana sonuçlar (3-5 madde)\n"
                    f"2. Üzerinde uzlaşılan noktalar\n"
                    f"3. Ayrışılan noktalar\n"
                    f"4. Somut aksiyon önerileri (5-10 madde)\n"
                    f"5. Genel değerlendirme"
                )},
            ],
        )
        summary = response["message"]["content"]
        self.key_ideas.append({
            "round": "final",
            "ideas": summary,
            "timestamp": datetime.now().isoformat(),
        })
        self.on_message("📊 Final Raporu", summary, 0)

    def _save_all(self):
        """Tüm sonuçları JSON, TXT ve Excel'e kaydet."""
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_topic = self.topic[:30].replace(" ", "_").replace("/", "-")
        base_name = f"{safe_topic}_{ts}"

        # JSON
        json_path = os.path.join(DEBATES_DIR, f"{base_name}.json")
        data = {
            "topic": self.topic,
            "personas": self.personas,
            "rounds": len(set(m["round"] for m in self.messages)),
            "total_messages": len(self.messages),
            "messages": self.messages,
            "key_ideas": self.key_ideas,
            "created": datetime.now().isoformat(),
        }
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        # TXT
        txt_path = os.path.join(DEBATES_DIR, f"{base_name}.txt")
        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(f"HAD3M-EIA TARTIŞMA RAPORU\n")
            f.write(f"{'=' * 50}\n")
            f.write(f"Konu: {self.topic}\n")
            f.write(f"Katılımcılar: {', '.join(self.personas)}\n")
            f.write(f"Tarih: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
            f.write(f"Toplam: {len(self.messages)} mesaj\n")
            f.write(f"{'=' * 50}\n\n")

            for m in self.messages:
                f.write(f"[Tur {m['round']}] {m['persona']}:\n{m['text']}\n\n")

            f.write(f"\n{'=' * 50}\n")
            f.write("ÖNEMLİ FİKİRLER VE ÇIKARIMLAR\n")
            f.write(f"{'=' * 50}\n\n")
            for k in self.key_ideas:
                f.write(f"--- Tur {k['round']} ---\n{k['ideas']}\n\n")

        # EXCEL
        xlsx_path = os.path.join(DEBATES_DIR, f"{base_name}.xlsx")
        self._save_excel(xlsx_path)

        self.on_status(f"Kaydedildi: {DEBATES_DIR}/{base_name}.*")

    def _save_excel(self, path: str):
        """Excel'e kaydet — mesajlar + fikirler + özet."""
        wb = Workbook()

        # Sheet 1: Tartışma
        ws = wb.active
        ws.title = "Tartışma"

        # Başlık
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

        headers = ["Tur", "Kişilik", "Mesaj", "Zaman"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for i, m in enumerate(self.messages, 2):
            ws.cell(row=i, column=1, value=m["round"])
            ws.cell(row=i, column=2, value=m["persona"])
            ws.cell(row=i, column=3, value=m["text"])
            ws.cell(row=i, column=3).alignment = Alignment(wrap_text=True)
            ws.cell(row=i, column=4, value=m["timestamp"][:19])

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 80
        ws.column_dimensions["D"].width = 20

        # Sheet 2: Fikirler & Özet
        ws2 = wb.create_sheet("Fikirler & Özet")
        ws2.cell(row=1, column=1, value="Tur").font = Font(bold=True)
        ws2.cell(row=1, column=2, value="Fikirler / Özet").font = Font(bold=True)

        for i, k in enumerate(self.key_ideas, 2):
            ws2.cell(row=i, column=1, value=str(k["round"]))
            ws2.cell(row=i, column=2, value=k["ideas"])
            ws2.cell(row=i, column=2).alignment = Alignment(wrap_text=True)

        ws2.column_dimensions["A"].width = 10
        ws2.column_dimensions["B"].width = 100

        # Sheet 3: Bilgi
        ws3 = wb.create_sheet("Bilgi")
        info = [
            ("Konu", self.topic),
            ("Katılımcılar", ", ".join(self.personas)),
            ("Toplam Tur", str(len(set(m["round"] for m in self.messages)))),
            ("Toplam Mesaj", str(len(self.messages))),
            ("Tarih", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Model", MODEL_NAME),
        ]
        for i, (k, v) in enumerate(info, 1):
            ws3.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws3.cell(row=i, column=2, value=v)

        ws3.column_dimensions["A"].width = 20
        ws3.column_dimensions["B"].width = 60

        wb.save(path)
